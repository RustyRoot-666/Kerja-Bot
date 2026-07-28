from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from services.order_repository import OrderRepository


ALIASES: dict[str, tuple[str, ...]] = {
    "ticket_id": (
        "TIKET ID", "TICKET ID", "TIKET", "TICKET", "INC", "NO TIKET",
    ),
    "service_number": (
        "NO SERVICE", "NO INET", "NO INTERNET", "INTERNET NUMBER",
        "SERVICE NUMBER", "INET", "NO_INET",
    ),
    "voip_number": (
        "NO VOIP", "VOIP", "NOMOR VOIP",
    ),
    "customer_name": (
        "NAMA", "NAMA PELANGGAN", "CUSTOMER NAME", "NAMA CUSTOMER",
        "CONTACT NAME",
    ),
    "address": (
        "ALAMAT", "ALAMAT PELANGGAN", "ADDRESS",
    ),
    "customer_phone": (
        "CP", "NO HP", "NO HP CUSTOMER", "CONTACT PHONE", "NOMOR HP",
    ),
    "old_sn": (
        "SN ONT LAMA", "SN LAMA", "SERIAL NUMBER LAMA", "SN OLD",
        "SN ONT OLD",
    ),
    "new_sn": (
        "SN ONT BARU", "SN BARU", "SERIAL NUMBER BARU", "SN NEW",
        "SN ONT NEW",
    ),
    "ont_type": (
        "TYPE ONT", "TIPE ONT", "MODEL ONT", "ONT TYPE", "GANTI KE",
        "GANTI ONT V. IBOOSTER",
    ),
    "sto": (
        "STO", "KODE STO", "ID STO", "ID_STO", "WORKZONE",
    ),
    "valins_id": (
        "VALINS ID", "VALIN ID", "VALINS", "VALIN",
    ),
    "result": (
        "RESULT", "HASIL", "STATUS HASIL", "STATUS",
    ),
    "config_description": (
        "KETERANGAN CONFIG", "KET CONFIG", "CONFIG DESCRIPTION",
    ),
    "report_description": (
        "KETERANGAN REPORT", "KET REPORT", "KETERANGAN STO",
        "REPORT DESCRIPTION", "KETERANGAN",
    ),
    "assigned_technician": (
        "NAMA PETUGAS", "PETUGAS", "TEKNISI", "NAMA TEKNISI",
        "ASSIGNED TECHNICIAN", "TECHNICIAN",
    ),
}


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().upper()
    text = re.sub(r"[\n\r\t]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def find_header_row(sheet, scan_rows: int = 20) -> tuple[int, dict[int, str]]:
    alias_lookup: dict[str, str] = {}
    for target, names in ALIASES.items():
        for name in names:
            alias_lookup[normalize_header(name)] = target

    best_row = 0
    best_mapping: dict[int, str] = {}
    max_row = sheet.max_row or 0

    if max_row == 0:
        raise ValueError("Sheet kosong.")

    for row_index in range(1, min(max_row, scan_rows) + 1):
        mapping: dict[int, str] = {}
        for column_index, cell in enumerate(sheet[row_index], start=1):
            normalized = normalize_header(cell.value)
            target = alias_lookup.get(normalized)
            if target:
                mapping[column_index] = target

        if len(mapping) > len(best_mapping):
            best_row = row_index
            best_mapping = mapping

    if not best_mapping:
        raise ValueError(
            "Header Excel tidak dikenali. Pastikan ada kolom TIKET/NO SERVICE/NAMA/ALAMAT."
        )

    return best_row, best_mapping


def select_order_sheets(workbook):
    """Utamakan sheet ORDER agar sheet bantuan/PIV/TACPRO tidak ikut terimpor."""
    for sheet_name in workbook.sheetnames:
        if normalize_header(sheet_name) == "ORDER":
            return [workbook[sheet_name]]
    return list(workbook.worksheets)


async def import_workbook(
    file_path: Path,
    repository: OrderRepository,
) -> dict[str, int]:
    workbook = load_workbook(file_path, read_only=False, data_only=True)

    stats = {
        "sheets": 0,
        "rows": 0,
        "inserted": 0,
        "updated": 0,
        "skipped": 0,
        "failed": 0,
    }

    try:
        for sheet in select_order_sheets(workbook):
            try:
                header_row, column_mapping = find_header_row(sheet)
            except ValueError:
                continue

            stats["sheets"] += 1
            max_row = sheet.max_row or header_row

            for row in sheet.iter_rows(
                min_row=header_row + 1,
                max_row=max_row,
                values_only=True,
            ):
                if not row or not any(value is not None for value in row):
                    continue

                stats["rows"] += 1
                data: dict[str, str] = {}

                for column_index, field_name in column_mapping.items():
                    if column_index - 1 < len(row):
                        data[field_name] = cell_text(row[column_index - 1])

                if not data.get("ticket_id") and not data.get("service_number"):
                    stats["skipped"] += 1
                    continue

                try:
                    result = await repository.upsert(data, source_file=file_path.name)
                    stats[result] += 1
                except Exception:
                    stats["failed"] += 1
    finally:
        workbook.close()

    if stats["sheets"] == 0:
        raise ValueError(
            "Tidak ada sheet yang memiliki header order yang dikenali."
        )

    return stats
